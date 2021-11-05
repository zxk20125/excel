import openpyxl
import os

files = {"input": "", "output": ""}
subject = []


def hashmap():
    data = {}
    for s in subject:
        data[s] = 0
    return data


def getpath():
    path = os.path.abspath(os.path.join(os.getcwd(), "../file"))
    for f in os.listdir(path):
        if f.endswith("target.xlsx"):
            files["input"] = path + '\\' + f
        elif f.endswith("output.xlsx"):
            files["output"] = path + '\\' + f


def get_subject():
    wb = openpyxl.load_workbook(files["input"])
    for name in wb.sheetnames:
        sheet = wb[name]
        for row in sheet.rows:
            for cell in row:
                if not subject.count(cell.value) and cell.value is not None:
                    subject.append(str(cell.value))
            break
    wb.close()


def write_subject():
    wb = openpyxl.load_workbook(files["output"])
    sheet = wb["Sheet2"]
    sheet.append(subject)
    wb.save(files["output"])


def init_data():
    tmp = {}
    for item in subject:
        tmp[item] = []
    return tmp


def get_data():
    sub = ""
    data = init_data()
    wb = openpyxl.load_workbook(files["input"])
    for sheet in wb.sheetnames:
        record = hashmap()
        ws = wb[sheet]
        for col in ws.columns:
            for cell in col:
                if subject.count(str(cell.value)):
                    sub = str(cell.value)
                elif sub != "" and cell.value != "":
                    data[sub].append(str(cell.value))
                    record[sub] = record[sub] + 1
            sub = ""
        for k, v in record.items():
            if v == 0:
                for i in range(record['项目']):
                    data[k].append("None")
    return data


def write_data():
    data = get_data()
    wo = openpyxl.load_workbook(files["output"])
    sheet2 = wo["Sheet1"]
    for sub in subject:
        # print(data[sub])
        sheet2.append(data[sub])
    wo.save(files["output"])


if __name__ == '__main__':
    getpath()
    get_subject()
    # write_subject()
    write_data()
